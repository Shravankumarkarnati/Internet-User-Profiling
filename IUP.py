# SHRAVAN KUMAR KARNATI

'''
Python Version 3.6.7 (default, Oct 22 2018, 11:32:17)
[GCC 8.2.0]
'''

# Modules/Libraries used
from __future__ import division

import datetime
import math
import os
import time

import pandas as pd
import scipy.stats
from openpyxl import load_workbook
from pandas import ExcelWriter

start_time = time.time()
dfW1 = pd.DataFrame()
dfW2 = pd.DataFrame()


def folder(array):
    # returns the first element of the array
    return array[0]


def epochConverter(epochTime):
    # converts the epoch time to the local time
    return (time.ctime(epochTime / 1000.0))  # Tue Feb  5 06:02:34 2013


def timeToSec(timeString):
    # converts local time to seconds
    x = time.strptime(timeString.split(',')[0], '%H:%M:%S')
    return (datetime.timedelta(hours=x.tm_hour, minutes=x.tm_min, seconds=x.tm_sec).total_seconds())


def convertSeconds(seconds):
    # converts seconds into HH:MM:SS format
    return str(time.strftime("%H:%M:%S", time.gmtime(seconds)))


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def avg(d, num):
    # finds the average doctets/Duration value for a day using all the Doctets/Duration values and time slot
    window = []
    av = []
    array = [x[0] for x in d]
    values = [x[1] for x in d]
    if num == 10:
        for i in range(28800, 61199 + 2, 10):
            window.append(i)
    if num == 227:
        for i in range(28800, 61199 + 2, 227):
            window.append(i)
    if num == 300:
        for i in range(28800, 61199 + 2, 300):
            window.append(i)
    while (len(window) > 1):
        x = 0
        y = 0
        for l in range(len(array)):
            if window[0] <= array[l] < window[1]:
                x = values[l] + x
                y += 1
        if y > 0 and x > 0:
            av.append(x / y)
        else:
            av.append(x)
        window.pop(0)
    return av


def findD(week, num, string, filepath):
    # Finds the average doctets/Duration value for a week data
    d = []
    d1 = []
    d2 = []
    d3 = []
    d4 = []
    d5 = []
    dateArray = list(dict.fromkeys(week['Date'].tolist()))

    for row in range(len(week.index)):
        if len(dateArray) > 0:
            if week['Date'][row] == dateArray[0]:
                d1.append((week['inSeconds'][row], week['d/D'][row]))
        if len(dateArray) > 1:
            if week['Date'][row] == dateArray[1]:
                d2.append((week['inSeconds'][row], week['d/D'][row]))
        if len(dateArray) > 2:
            if week['Date'][row] == dateArray[2]:
                d3.append((week['inSeconds'][row], week['d/D'][row]))
        if len(dateArray) > 3:
            if week['Date'][row] == dateArray[3]:
                d4.append((week['inSeconds'][row], week['d/D'][row]))
        if len(dateArray) > 4:
            if week['Date'][row] == dateArray[4]:
                d5.append((week['inSeconds'][row], week['d/D'][row]))

    d1 = avg(d1, num)
    d2 = avg(d2, num)
    d3 = avg(d3, num)
    d4 = avg(d4, num)
    d5 = avg(d5, num)

    d = d1 + d2 + d3 + d4 + d5
    if string == "Week1":
        dfW1[string] = d
    else:
        dfW2[string] = d


def createWeek(filepath, max_row, week, timeSlot):
    # creates week windows from excel files into a dataframe / returns a dataframe
    dfS1 = pd.ExcelFile(filepath).parse('Sheet1')
    week1_start = 1359982800000  # Feb 4 01:00:00 GMT
    week1_end = 1360360799599  # Feb 8 09:59:59 GMT
    week2_start = 1360587600000  # Feb 11 01:00:00 GMT
    week2_end = 1360965599599  # Feb 15 09:59:59 GMT
    y = []
    z = []
    x = []
    m = []
    Days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    day = 0
    date = 0
    if week == "Week1":
        for row in range(0, max_row):
            if week1_start <= dfS1['Real First Packet'][row] <= week1_end and (dfS1['Duration'][row]) > 0:
                dateString = epochConverter(dfS1['Real First Packet'][row])
                if "08:00:00" <= (dateString.split()[3]) <= "16:59:59":
                    for n in range(0, len(Days)):
                        if Days[n] == (dateString.split()[0]):
                            x.append(dateString.split()[2])
                            m.append(dateString)
                            y.append((dfS1['doctets'][row]) / (dfS1['Duration'][row]))
                            z.append(timeToSec(dateString.split()[3]))

    elif week == "Week2":
        for row in range(0, max_row):
            if week2_start <= dfS1['Real First Packet'][row] <= week2_end and (dfS1['Duration'][row]) > 0:
                dateString = epochConverter(dfS1['Real First Packet'][row])
                if "08:00:00" <= (dateString.split()[3]) <= "16:59:59":
                    for n in range(0, len(Days)):
                        if Days[n] == (dateString.split()[0]):
                            x.append(dateString.split()[2])
                            m.append(dateString)
                            y.append((dfS1['doctets'][row]) / (dfS1['Duration'][row]))
                            z.append(timeToSec(dateString.split()[3]))

    dfX = pd.DataFrame()
    dfX['Date'] = x
    dfX['dateString'] = m
    dfX['d/D'] = y
    dfX['inSeconds'] = z
    findD(dfX, timeSlot, week, filepath)


def spearmans_rank_correlation(xs, ys):
    # Calculate the rank of x's
    xranks = pd.Series(xs).rank()

    # Caclulate the ranking of the y's
    yranks = pd.Series(ys).rank()

    # Calculate Pearson's correlation coefficient on the ranked versions of the data
    return scipy.stats.pearsonr(xranks, yranks)


def points(a):
    # If the rank is equal to 1, changes it to 0.99
    if a == 1:
        return 0.99
    elif math.isnan(a):
        return 0
    else:
        return a


def findZ(r1a2a, r1a2b, r2a2b, timeSlot):
    # Finding the Z value using the ranks
    N = 0
    rmsq2 = 0
    f = 0
    h = 0
    Z2 = 0
    Z1 = 0
    p = 0
    q = 0
    N = (61200 - 28800) * 5 / timeSlot
    rmsq2 = (r1a2a * r1a2a + r1a2b * r1a2b) / 2
    f = (1 - r2a2b) / 2 * (1 - rmsq2)
    h = (1 - (f * rmsq2)) / (1 - rmsq2)
    Z2 = 0.5 * math.log((1 + r1a2b) / (1 - r1a2b))
    Z1 = 0.5 * math.log((1 + r1a2a) / (1 - r1a2a))
    p = Z1 - Z2
    q = math.sqrt((N - 3)) / 2 * (1 - r2a2b) * h

    value = p * q
    return value


def findP(z):
    # finds the final P value using the Z value
    p = 0.3275911
    a1 = 0.254829592
    a2 = -0.284496736
    a3 = 1.421413741
    a4 = -1.453152027
    a5 = 1.061405429
    sign = 0
    if (z < 0.0):
        sign = -1
    else:
        sign = 1

    x = abs(z) / math.sqrt(2.0)
    t = (1.0) / (1.0 + p * x)
    erf = 1.0 - (((((a5 * t + a4) * t) + a3) * t + a2) * t + a1) * t * math.exp(-x * x)
    phiZ = 0.5 * (1.0 + sign * erf)
    return 1 - phiZ


def extract(filepath1, filepath2, Sheet1, Sheet2):
    # Function to extract the week's data and saving the values into a list
    df1w1 = pd.ExcelFile(filepath1).parse(Sheet1)
    df1w2 = pd.ExcelFile(filepath1).parse(Sheet2)
    df2w1 = pd.ExcelFile(filepath2).parse(Sheet1)
    df2w2 = pd.ExcelFile(filepath2).parse(Sheet2)

    # aw1=[],aw2=[],bw1=[],bw2=[]
    aw1 = df1w1['Week1'].tolist()
    aw2 = df1w2['Week2'].tolist()
    bw1 = df2w1['Week1'].tolist()
    bw2 = df2w2['Week2'].tolist()
    return aw1, aw2, bw1, bw2


# Path to the excel Files of all users
dirpath = r"C:\Users\shrav\Desktop\InfoSec Excels"
a = []
finalValues = []
dist = 0
inDist = 0

# timeSlot(Please change it to what ever value you want to calculate for)
timeSlot = 300

if timeSlot == 10:
    Sheet1 = "Sheet2"
    Sheet2 = "Sheet3"
elif timeSlot == 227:
    Sheet1 = "Sheet4"
    Sheet2 = "Sheet5"
elif timeSlot == 300:
    Sheet1 = "Sheet6"
    Sheet2 = "Sheet7"

# Saves all the file values into a text file.
file1 = open(r"C:\Users\shrav\Desktop\InfoSec Excels\values.txt", "a+")
for filename in os.listdir(dirpath):
    if filename.endswith("xlsx"):
        a.append(filename)
print(a)

# All the possible combination for the users
b = [(x, y) for x in a for y in a]
df = pd.DataFrame(columns=a, index=a)
i = 0

for file in range(len(a)):
    '''
    Pre Processing

    This loop takes all the excel files and only takes the two weeks data with exceptions
    of leaving saturday and sunday, and the entries that are between 8 am to 5 pm.

    Then saves these values into the same excel files as Sheets.  
    '''
    start_time1 = time.time()
    filename = folder(a)
    print(len(a), filename)
    filepath = (os.path.join(dirpath, filename))

    # for file1
    # load file.xlsx
    wb1 = load_workbook(filepath)

    # select file.xlsx
    sheet1 = wb1.active

    # get max row count
    max_row1 = sheet1.max_row - 1

    (createWeek(filepath, max_row1, "Week1", timeSlot))

    (createWeek(filepath, max_row1, "Week2", timeSlot))

    append_df_to_excel(filepath, dfW1, sheet_name=Sheet1, startrow=0)
    append_df_to_excel(filepath, dfW2, sheet_name=Sheet2, startrow=0)

    a.pop(0)
    print("doctets/Duration values for Week1 and Week 2 Insertion done for " + filename)
    print("--- A total of %s minutes taken ---" % ((time.time() - start_time1) / 60))

for k in range(len(b)):
    '''
    Calculation 

    In this loop, the values of P for all the possible combinations is calucated for the given timeslot.

    '''
    i += 1
    start_time2 = time.time()
    filename1 = folder(b)[0]
    filename2 = folder(b)[1]
    filepath1 = (os.path.join(dirpath, filename1))
    filepath2 = (os.path.join(dirpath, filename2))

    (aw1, aw2, bw1, bw2) = extract(filepath1, filepath2, Sheet1, Sheet2)

    r1a2a = (spearmans_rank_correlation(aw1, aw2)[0])

    r1a2b = (spearmans_rank_correlation(aw1, bw2)[0])

    r2a2b = (spearmans_rank_correlation(aw2, bw2)[0])

    r1a2a = points(r1a2a)

    r1a2b = points(r1a2b)

    r2a2b = points(r2a2b)

    # print(r1a2a, r1a2b, r2a2b)

    z = findZ(r1a2a, r1a2b, r2a2b, timeSlot)

    p = findP(z)
    # prints the P value.
    print(p)
    if timeSlot:
        file1.write(filename1)
        file1.write("   ")
        file1.write(filename2)
        file1.write("   ")
        file1.write(str(p))
        file1.write("\n")
        df.at[filename1, filename2] = p
        finalValues.append(p)
        if p <= 0.05:
            dist += 1
        else:
            inDist += 1

    print("One Combination Completed!", filename1, filename2, i)
    print("--- %s minutes taken ---" % ((time.time() - start_time2) / 60))
    print("-" * 40)
    b.pop(0)

file1.close()
writer = ExcelWriter('10 Seconds Final.xlsx')
df.to_excel(writer, 'Sheet1')
writer.save()

print("Distinguishable are:  %s" % dist)
print("Indistinguishable are:  %s" % inDist)
print("Completed!")
print("--- A total of %s minutes taken ---" % ((time.time() - start_time) / 60))
